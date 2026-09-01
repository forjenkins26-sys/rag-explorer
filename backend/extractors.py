import csv
import io
import re
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader

SUPPORTED_EXTENSIONS = {".pdf", ".txt", ".md", ".xlsx", ".xls", ".csv"}


def _normalize(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def _extract_pdf(path: Path) -> list[tuple[str, str]]:
    reader = PdfReader(str(path))
    return [
        (str(i + 1), _normalize(page.extract_text() or ""))
        for i, page in enumerate(reader.pages)
    ]


def _extract_text_file(path: Path) -> list[tuple[str, str]]:
    raw = path.read_text(encoding="utf-8", errors="ignore")
    return [("1", _normalize(raw))]


def _extract_xlsx(path: Path) -> list[tuple[str, str]]:
    wb = load_workbook(str(path), data_only=True, read_only=True)
    sections = []
    for sheet in wb.worksheets:
        rows_text = []
        for row in sheet.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None]
            if cells:
                rows_text.append(" | ".join(cells))
        if rows_text:
            sections.append((sheet.title, _normalize("\n".join(rows_text))))
    return sections


def _extract_csv(path: Path) -> list[tuple[str, str]]:
    """One section per data row, labelled by its first column.

    Deliberately NOT one section for the whole file. A CSV is usually a table of
    discrete records - test cases, tickets, products - and each row is the unit a
    question is really about. Emitting rows separately keeps a retrieved chunk
    from straddling two unrelated records.

    Each row is rendered as "Header: value" pairs so the embedding carries the
    column meaning, not just the bare value. `_normalize` is applied per FIELD
    rather than to the whole row: it collapses all whitespace, which would erase
    the newlines that separate one field from the next.
    """
    raw = path.read_text(encoding="utf-8-sig", errors="ignore")

    # Sniff the delimiter so semicolon and tab files work too. Fall back to a
    # comma when the sample is too uniform for the sniffer to decide.
    try:
        dialect = csv.Sniffer().sniff(raw[:4096], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel

    rows = list(csv.reader(io.StringIO(raw), dialect))
    if not rows:
        return []

    header = [(_normalize(h) or f"column{i + 1}") for i, h in enumerate(rows[0])]
    sections: list[tuple[str, str]] = []

    for line_number, row in enumerate(rows[1:], start=2):
        if not any((cell or "").strip() for cell in row):
            continue  # blank line

        parts = []
        for index, cell in enumerate(row):
            value = _normalize(cell or "")
            if not value:
                continue
            name = header[index] if index < len(header) else f"column{index + 1}"
            parts.append(f"{name}: {value}")

        if not parts:
            continue

        # Label the section by the first column's value when it looks like an id
        # (LOGIN-001, SCRUM-255); otherwise fall back to the line number.
        first = _normalize(row[0] or "")
        label = first if first and len(first) <= 40 else f"row {line_number}"
        sections.append((label, "\n".join(parts)))

    return sections


_EXTRACTORS = {
    ".pdf": _extract_pdf,
    ".txt": _extract_text_file,
    ".md": _extract_text_file,
    ".xlsx": _extract_xlsx,
    ".xls": _extract_xlsx,
    ".csv": _extract_csv,
}


def extract_sections(path: Path) -> list[tuple[str, str]]:
    """Returns [(location_label, text), ...] — page number for PDF, sheet name for
    Excel, "1" for flat text files. Raises ValueError for unsupported extensions."""
    ext = path.suffix.lower()
    extractor = _EXTRACTORS.get(ext)
    if extractor is None:
        raise ValueError(f"Unsupported file type: {ext}")
    return extractor(path)
